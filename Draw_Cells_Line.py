#!/usr/bin/env python
#-*- coding=utf-8 -*-

import os
import os.path
import string
import logging
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl.styles as sts
import openpyxl.cell as ce



def style_range(ws, cell_range, style=None):
    """
    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param style: An openpyxl Style object
    """

    start_cell, end_cell = cell_range.split(':')
    start_coord = ce.coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = ce.column_index_from_string(start_coord[0])
    end_coord = ce.coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = ce.column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            col = ce.get_column_letter(col_idx)
            ws.cell('%s%s' % (col, row)).style = style   

def get_all_cells(ws,cell_ranges):
        """
        根据给定的区域，获取该区域中的所有单元格 A4:A5
        """
        result_cells = []
        min_col = 0
        min_row = 0
        max_col = 0
        max_row = 0
        cell_start,cell_end = cell_ranges.split(":")
        min_col,min_row = ce.column_index_from_string(ws[cell_start].column),ws[cell_start].row
        max_col,max_row = ce.column_index_from_string(ws[cell_end].column),ws[cell_end].row
        for c_t in range(min_col,max_col+1):
            for r_t in range(min_row,max_row+1):
                result_cells.append(ce.get_column_letter(c_t)+str(r_t))
        return result_cells
    
    
def get_merged_range(ws):
        """
        获取表格中所有合并单元格的区域
        """
        merged_ranges = []
        all_merged_ranges = ws.merged_cell_ranges
        for tmp in all_merged_ranges:
            merged_ranges.append(tmp)
        return merged_ranges


def Draw_Cells_Line(ws):
    """
    补全合并单元格的边框线条
    """
    merged_ranges = get_merged_range(ws)
    for merged_range in merged_ranges:
        cells = get_all_cells(ws, merged_range)
        start_cell = ws.cell(cells[0])
        top_line_style = start_cell.style.border.top.style
        bot_line_style = start_cell.style.border.bottom.style
        left_line_style = start_cell.style.border.left.style
        
        if ws.cell(cells[0]).column == ws.cell(cells[-1]).column:
            #同一列
            style_range(ws, merged_range, sts.Style(
                               border=sts.Border(top=sts.Side(border_style=top_line_style, color=sts.colors.BLACK),
                                             left=sts.Side(border_style=left_line_style, color=sts.colors.BLACK),
                                             bottom=sts.Side(border_style=bot_line_style, color=sts.colors.BLACK),
                                             right=sts.Side(border_style=left_line_style, color=sts.colors.BLACK), ), alignment=sts.alignment.Alignment(horizontal='center', vertical='center')),)
        elif ws.cell(cells[0]).row == ws.cell(cells[-1]).row:
            #同一行
            style_range(ws, merged_range, sts.Style(
                               border=sts.Border(top=sts.Side(border_style=top_line_style, color=sts.colors.BLACK),
                                             left=sts.Side(border_style=left_line_style, color=sts.colors.BLACK),
                                             bottom=sts.Side(border_style=top_line_style, color=sts.colors.BLACK),
                                             right=sts.Side(border_style=left_line_style, color=sts.colors.BLACK), ), alignment=sts.alignment.Alignment(horizontal='center', vertical='center')),)
    return ws
   

       
       
if __name__ == "__main__":
    wb = load_workbook("001.xlsx")
    ws = wb.active
    wss = Draw_Cells_Line(ws)    
    wb.save("saved.xlsx")
    print "End!"    
            
    
    